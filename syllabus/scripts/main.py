import openpyxl
import json
from collections import OrderedDict
from tqdm import tqdm
import argparse

# 辞書を生成するときに使用する関数

def generation_dict(config_path, key):
    f = open(config_path, 'r')
    dict_name = f.readlines()
    f.close()
    if key <= 20 or key >= 83:
        return dict_name[key].rstrip('\n')
    elif key % 2 != 0:
        return "theme"
    elif key % 2 == 0:
        return "homework"


def convert_json(import_path, export_path, config_path):
    excel_path = import_path
    workbook = openpyxl.load_workbook(filename=excel_path, read_only=False)
    # シートのロード
    sheet = workbook['Sheet0']

    json_list = []
    first_counter = True
    # tqdm(sheet.iter_rows(), total=len(len(sheet.iter_rows())))
    # sheet.iter_rows()
    print(sheet.max_row, sheet.max_column)
    sheet_len = sheet.max_row
    for row in tqdm(sheet.iter_rows(), total=sheet_len):
        theme = []
        homework = []

        if first_counter:
            first_counter = False
            continue
        else:
            data = OrderedDict()
        for cell in row:
            if cell.row >= 2:
                value = str(cell.value).replace('\n', '')
                if generation_dict(config_path, int(cell.column) - 1) == "theme":
                    theme.append(value)
                    data[generation_dict(config_path, int(cell.column) - 1)] = theme
                elif generation_dict(config_path, int(cell.column) - 1) == "homework":
                    homework.append(value)
                    data[generation_dict(config_path, int(cell.column) - 1)] = homework
                else:
                    data[generation_dict(config_path, int(cell.column) - 1)] = value
                
        json_list.append(data)

    with open(export_path, 'w') as file:
        json.dump(json_list, file, indent=4, ensure_ascii=False)

    print("Success!!")


def main():
    option = argparse.ArgumentParser(description='スクリプトのオプション')
    option.add_argument('--import_path', type=str, default='syllabus.xlsx', help='.xlsx import file path')
    option.add_argument('--export_path', type=str, default='output.json', help='output json file path')
    option.add_argument('--config_path', type=str, default='dict_name.txt', help='dict config file path')
    args = option.parse_args()

    convert_json(args.import_path, args.export_path, args.config_path)


if __name__ == "__main__":
    main()

